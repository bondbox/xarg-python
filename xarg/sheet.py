# coding=utf-8

from csv import DictReader as csv_dist_reader
from csv import DictWriter as csv_dist_writer
from csv import reader as csv_reader
from csv import writer as csv_writer
import os
from typing import Any
from typing import Callable
from typing import Dict
from typing import Generic
from typing import Iterable
from typing import Iterator
from typing import List
from typing import Optional
from typing import Tuple
from typing import TypeVar
from typing import Union

import openpyxl
from tabulate import TableFormat
from tabulate import tabulate as __tabulate
from wcwidth import wcswidth
import xlrd
import xlwt

FKT = TypeVar("FKT")
FVT = TypeVar("FVT")
RKT = TypeVar("RKT")
RVT = TypeVar("RVT")
CVT = TypeVar("CVT")


class cell(Generic[CVT]):
    """Cell in the custom table

    Define cells to resolve the null value issue.
    """

    def __init__(self, value: Optional[CVT] = None):
        self.value = value

    def __str__(self) -> str:
        return str(self.value if self.value is not None else "")

    @property
    def empty(self) -> bool:
        return self.value is None

    @property
    def value(self) -> Optional[CVT]:
        return self.__value

    @value.setter
    def value(self, value: Optional[CVT]):
        self.__value = value


class row(Generic[RKT, RVT]):
    """Row in the custom table
    """

    def __init__(self, values: Union[Iterable[cell[RVT]],
                                     Iterable[Optional[RVT]]]):
        self.__cells = [self.new_cell(value=value) for value in values]

    def __len__(self) -> int:
        return len(self.__cells)

    def __iter__(self) -> Iterator[cell[RVT]]:
        """all cells
        """
        return iter(self.__cells)

    def __getitem__(self, index: int) -> cell[RVT]:
        return self.__cells[index]

    def __setitem__(self, index: int,
                    value: Union[cell[RVT],
                                 Optional[RVT]]
                    ) -> None:
        self.__cells[index] = self.new_cell(value=value)

    @property
    def values(self) -> Tuple[Optional[RVT], ...]:
        """all cell values
        """
        return tuple(cell.value for cell in self)

    def append(self, value: Union[cell[RVT], Optional[RVT]]) -> None:
        self.__cells.append(self.new_cell(value))

    def extend(self, values: Union[Iterable[cell[RVT]],
                                   Iterable[Optional[RVT]]]) -> None:
        self.__cells.extend(self.new_cell(value) for value in values)

    def mapping(self, header: Tuple[RKT, ...]) -> Dict[RKT, RVT]:
        """Map the value of cells into dict
        """
        return {key: cell.value for key, cell in zip(header, self)
                if cell.value is not None}

    def new_cell(self, value: Union[cell[RVT], Optional[RVT]]) -> cell[RVT]:
        return value if isinstance(value, cell) else cell(value)


class form(Generic[FKT, FVT]):
    """Custom table
    """

    def __init__(self, name: str, header: Optional[Iterable[FKT]] = None):
        self.__rows: List[row[FKT, FVT]] = []
        self.__name: str = name
        self.header = header if header is not None else []

    def __len__(self) -> int:
        return len(self.__rows)

    def __iter__(self) -> Iterator[row[FKT, FVT]]:
        """all rows
        """
        return iter(self.__rows)

    def __getitem__(self, index: int) -> row[FKT, FVT]:
        return self.__rows[index]

    def __setitem__(self, index: int,
                    value: Union[row[FKT, FVT],
                                 Iterable[cell[FVT]],
                                 Iterable[FVT]]
                    ) -> None:
        self.__rows[index] = self.new_row(value)

    @property
    def name(self) -> str:
        """table name
        """
        return self.__name

    @property
    def header(self) -> Tuple[FKT, ...]:
        """table header (title line)
        """
        return self.__header

    @header.setter
    def header(self, value: Iterable[FKT]) -> None:
        self.__header: Tuple[FKT, ...] = tuple(i for i in value)

    @property
    def mappings(self) -> Iterator[Dict[FKT, FVT]]:
        return iter(row.mapping(self.header) for row in self)

    @property
    def values(self) -> Tuple[Tuple[Optional[FVT], ...], ...]:
        """all cell values (by row)
        """
        return tuple(row.values for row in self)

    def column_no(self, key: FKT) -> int:
        return self.header.index(key)

    def sort(self, key: Callable[[row[FKT, FVT]], cell[FVT]],
             reverse: bool = False) -> None:
        """sort rows using a Lambda function as the key.
        """
        self.__rows.sort(key=lambda row: key(row).value,  # type: ignore
                         reverse=reverse)

    def dump(self) -> Tuple[Tuple[Any, ...], ...]:
        """dump header and all rows
        """
        table: List[Tuple[Any, ...]] = [self.header]
        table.extend(self.values)
        return tuple(table)

    def reflection(self, cells: Dict[FKT, FVT],
                   default: Any = None) -> row[FKT, FVT]:
        """Re-map the dict to new row object
        """
        return self.new_row(cells=tuple(cells.get(key, default)
                                        for key in self.header))

    def append(self, item: Union[row[FKT, FVT],
                                 Iterable[cell[FVT]],
                                 Iterable[FVT]]
               ) -> None:
        self.__rows.append(self.new_row(item))

    def extend(self, rows: Iterable[Union[row[FKT, FVT],
                                          Iterable[cell[FVT]],
                                          Iterable[FVT]]]) -> None:
        self.__rows.extend(self.new_row(row) for row in rows)

    def new_row(self, cells: Union[row[FKT, FVT],
                                   Iterable[cell[FVT]],
                                   Iterable[Optional[FVT]]]
                ) -> row[FKT, FVT]:
        return cells if isinstance(cells, row) else row(values=cells)

    def new_map(self, default: Any = None) -> Dict[FKT, Any]:
        """Generate new mapping with default values
        """
        return {key: default for key in self.header}


def tabulate(table: form[Any, Any],
             fmt: Union[str, TableFormat] = "simple") -> str:
    return __tabulate(tabular_data=table.values,
                      headers=table.header,
                      tablefmt=fmt)


def parse_table_name(filename: str) -> str:
    return os.path.splitext(os.path.basename(filename))[0]


class csv():

    @classmethod
    def load(cls, filename: str,
             include_header: bool = True
             ) -> form[str, str]:
        """Read .csv file
        """
        with open(filename, "r", encoding="utf-8") as rhdl:
            table: form[str, str] = form(name=parse_table_name(filename))
            if include_header:
                reader = csv_dist_reader(rhdl)
                fields = reader.fieldnames
                if fields is not None:
                    table.header = fields
                    for _row in reader:
                        table.append(table.reflection(_row))
            else:
                reader = csv_reader(rhdl)
                for _row in reader:
                    table.append(_row)
        return table

    @classmethod
    def dump(cls, filename: str, table: form[Any, Any]) -> None:
        """Write .csv file
        """
        with open(filename, "w", encoding="utf-8") as whdl:
            if len(table.header) > 0:
                writer = csv_dist_writer(whdl, fieldnames=table.header)
                writer.writeheader()
                writer.writerows(table.mappings)
            else:
                writer = csv_writer(whdl)
                writer.writerows(table.dump())


class xls_reader():
    """Read .xls file
    """

    def __init__(self, filename: str):
        self.__book: xlrd.Book = xlrd.open_workbook(filename)
        self.__file: str = filename

    @property
    def file(self) -> str:
        return self.__file

    @property
    def book(self) -> xlrd.Book:
        return self.__book

    def load_sheet(self, sheet_name: Optional[str] = None) -> form[str, str]:
        sheet_index: int = self.book.sheet_names().index(sheet_name)\
            if isinstance(sheet_name, str) else 0
        sheet: xlrd.sheet.Sheet = self.book.sheet_by_index(sheet_index)
        first: Iterable[str] = sheet.row_values(0)  # first line as header
        table: form[str, Any] = form(name=sheet.name, header=first)
        for i in range(1, sheet.nrows):
            table.append(sheet.row_values(i))
        return table

    def load_sheets(self) -> Tuple[form[str, str], ...]:
        return tuple(self.load_sheet(name) for name in self.book.sheet_names())


class xls_writer():
    """Write .xls file
    """
    WIDTH = 325

    def __init__(self):
        self.__book: xlwt.Workbook = xlwt.Workbook(
            encoding="utf-8", style_compression=0)

    @property
    def book(self) -> xlwt.Workbook:
        return self.__book

    def save(self, filename: str) -> bool:
        abspath: str = os.path.abspath(filename)
        try:
            dirname: str = os.path.dirname(abspath)
            if not os.path.exists(dirname):
                os.makedirs(dirname)
            self.book.save(abspath)
            return True
        except Exception:  # pylint: disable=broad-except
            # f"failed to write file {abspath}"
            return False

    def dump_sheet(self, table: form[Any, Any]):
        sheet: xlwt.Worksheet = self.book.add_sheet(
            table.name, cell_overwrite_ok=True)
        widths: List[int] = []
        values: Tuple[Tuple[Any, ...], ...] = table.dump()
        for row_no, cells in enumerate(values):
            for col_no, _cell in enumerate(cells):
                value = str(_cell)
                sheet.write(row_no, col_no, value)
                width = wcswidth(value)
                if col_no >= len(widths):
                    sheet.col(col_no).width = self.WIDTH
                    widths.append(0)
                if width > widths[col_no]:
                    sheet.col(col_no).width = self.WIDTH * width
                    widths[col_no] = width

    def dump_sheets(self, tables: Iterable[form[Any, Any]]):
        for table in tables:
            self.dump_sheet(table=table)


class xlsx():
    """Read or write .xlsx file
    """

    def __init__(self, filename: str, read_only: bool = True):
        self.__book: openpyxl.Workbook = openpyxl.load_workbook(
            filename=filename, read_only=read_only)
        self.__file: str = filename

    @property
    def file(self) -> str:
        return self.__file

    @property
    def book(self) -> openpyxl.Workbook:
        return self.__book

    def load_sheet(self, sheet_name: Optional[str] = None) -> form[str, Any]:
        def get_default_sheet_name() -> str:
            if isinstance(sheet_name, str):
                return sheet_name
            active_sheet = self.book.active
            if active_sheet is not None:
                return active_sheet.title
            return self.book.sheetnames[0]

        sheet = self.book[get_default_sheet_name()]
        first = list(sheet.iter_rows(max_row=1))[0]
        cells: List[str] = [c.value for c in first if isinstance(c.value, str)]
        table: form[str, Any] = form(name=sheet.title, header=cells)
        for _row in sheet.iter_rows(min_row=2):
            table.append([cell.value for cell in _row])
        return table

    def load_sheets(self) -> Tuple[form[str, str], ...]:
        return tuple(self.load_sheet(name) for name in self.book.sheetnames)
